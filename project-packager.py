##MBTA Workspace GID: 15492006741476
##Test bed project GID : 1203186680032258
logo_str = """ 
█████╗ ██╗   ██╗████████╗ ██████╗     ██████╗  █████╗  ██████╗██╗  ██╗ █████╗  ██████╗ ███████╗██████╗ 
██╔══██╗██║   ██║╚══██╔══╝██╔═══██╗    ██╔══██╗██╔══██╗██╔════╝██║ ██╔╝██╔══██╗██╔════╝ ██╔════╝██╔══██╗
███████║██║   ██║   ██║   ██║   ██║    ██████╔╝███████║██║     █████╔╝ ███████║██║  ███╗█████╗  ██████╔╝
██╔══██║██║   ██║   ██║   ██║   ██║    ██╔═══╝ ██╔══██║██║     ██╔═██╗ ██╔══██║██║   ██║██╔══╝  ██╔══██╗
██║  ██║╚██████╔╝   ██║   ╚██████╔╝    ██║     ██║  ██║╚██████╗██║  ██╗██║  ██║╚██████╔╝███████╗██║  ██║
╚═╝  ╚═╝ ╚═════╝    ╚═╝    ╚═════╝     ╚═╝     ╚═╝  ╚═╝ ╚═════╝╚═╝  ╚═╝╚═╝  ╚═╝ ╚═════╝ ╚══════╝╚═╝  ╚═╝
"""
import sys
import os
import pandas
import asana
import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import date
##TODO: Actually look into those deprecation warnings
##TODO: Standardize your capitalization and how you format functions
###FUNCTIONS###
##TODO: ALPHABETIZE TEST TASKS AFTER GENERAL COMMENTS -- DONE I THINK
##TODO: Delete unecessary tabs after tasks are done.
def AddTestTasksToWorksheet(gen_tasks, test_tasks, sheet, workbook):
##QA COMPLETE TAG GID = 649069647070258
##HOLD TAG GID = 259956811260129
##STATUS COLUMN HEADER = OPYXL ROW 10, COLUMN 10 
    qa_tag = {'gid': '649069647070258', 'resource_type': 'tag'}
    hold_tag =  {'gid': '259956811260129', 'resource_type': 'tag'}
    rownum = 11
    ##TODO RIGHT HERE: split off anything without 'ID Number' into its own list. Sort second list by name alphabetically. Remerge.
    for row, task in enumerate(gen_tasks):
            print(task['name'])
        ##OPYXL is 1-indexed. This should be the first row of the output spreadsheet you want to fill. TODO: Find the actual header and drop down one instead of writing '11'
        ##TODO BELOW: If ID number is missing and if there is a Design Review Comment Status, append that in column 10 instead of Test Script Review Status
            if hold_tag not in task['tags']:
                if qa_tag in task['tags']:
                    for ind, field in enumerate(task['custom_fields']):

                        if (field['name'] == 'Test Script Review Status'):
                            sheet.cell(row = rownum, column = 10).value = task['custom_fields'][ind]['display_value']
                        ##Design Review Overwrites Test Script Review Status if it exists
                        if(field['name'] == 'Design Review Comment Status'):
                            if(task['custom_fields'][ind]['display_value'] != None):
                                sheet.cell(row = rownum, column = 10).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'ID Number'):
                            sheet.cell(row = rownum, column = 5).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                            ##Test Type overrules comment tracking if present
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Functional area'):
                            sheet.cell(row = rownum, column = 4).value = task['custom_fields'][ind]['display_value']
                        ##If no functional area is defined, add the task section.
                            if (task['custom_fields'][ind]['display_value'] == None):
                                sheet.cell(row = rownum, column = 4).value = task['section']
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 3).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']

                    sheet.cell(row = rownum, column = 7).value = task['notes']
                    sheet.cell(row = rownum, column = 6).value = task['name']
                    rownum = rownum + 1
    for row, task in enumerate(test_tasks):
            print(task['name'])
        ##OPYXL is 1-indexed. This should be the first row of the output spreadsheet you want to fill. TODO: Find the actual header and drop down one instead of writing '11'
        ##TODO BELOW: If ID number is missing and if there is a Design Review Comment Status, append that in column 10 instead of Test Script Review Status
            if hold_tag not in task['tags']:
                if qa_tag in task['tags']:
                    for ind, field in enumerate(task['custom_fields']):

                        if (field['name'] == 'Test Script Review Status'):
                            sheet.cell(row = rownum, column = 10).value = task['custom_fields'][ind]['display_value']
                        ##Design Review Overwrites Test Script Review Status if it exists
                        if(field['name'] == 'Design Review Comment Status'):
                            if(task['custom_fields'][ind]['display_value'] != None):
                                sheet.cell(row = rownum, column = 10).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'ID Number'):
                            sheet.cell(row = rownum, column = 5).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                            ##Test Type overrules comment tracking if present
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Functional area'):
                            sheet.cell(row = rownum, column = 4).value = task['custom_fields'][ind]['display_value']
                        ##If no functional area is defined, add the task section.
                            if (task['custom_fields'][ind]['display_value'] == None):
                                sheet.cell(row = rownum, column = 4).value = task['section']
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 3).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Test Type'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']

                    sheet.cell(row = rownum, column = 7).value = task['notes']
                    sheet.cell(row = rownum, column = 6).value = task['name']
                    rownum = rownum + 1
                        
            workbook.save('output.xlsm')

def AddStandardTasksToWorksheetNoLinks(tasks, sheet, workbook):
##QA COMPLETE TAG GID = 649069647070258
##HOLD TAG GID = 259956811260129
##STATUS COLUMN HEADER = OPYXL ROW 10, COLUMN 10 
    qa_tag = {'gid': '649069647070258', 'resource_type': 'tag'}
    hold_tag =  {'gid': '259956811260129', 'resource_type': 'tag'}
    rownum = 11
    for row, task in enumerate(tasks):
        ##OPYXL is 1-indexed. This should be the first row of the output spreadsheet you want to fill. TODO: Find the actual header and drop down one instead of writing '11'
            print(task['name'])
            if hold_tag not in task['tags']:
                if qa_tag in task['tags']:
                    for ind, field in enumerate(task['custom_fields']):
                        print(field['name'])
                        ##Add review status...
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Submittal ID (CDRL number and section)'):
                            sheet.cell(row = rownum, column = 5).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'MBTA Owner'):
                            sheet.cell(row = rownum, column = 3).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Design Review Comment Status'):
                            sheet.cell(row = rownum, column = 11).value = task['custom_fields'][ind]['display_value']
                    sheet.cell(row = rownum, column = 7).value = task['notes']
                    sheet.cell(row = rownum, column = 6).value = task['name']
                    sheet.cell(row = rownum, column = 4).value = task['section']
                    time = task['created_at']
                    split_time = time.split('T')[0]
                    strip_time = datetime.datetime.strptime(split_time, "%Y-%m-%d")
                    time_str = strip_time.strftime('%m/%d/%Y')
                    print(time_str)

                    sheet.cell(row = rownum, column = 8).value = time_str
                    rownum = rownum + 1
            
            workbook.save('output.xlsm')
##TODO: Finish with-links version
def AddStandardTasksToWorksheetWithLinks(tasks, sheet, workbook):
##QA COMPLETE TAG GID = 649069647070258
##HOLD TAG GID = 259956811260129
##STATUS COLUMN HEADER = OPYXL ROW 10, COLUMN 10 
    qa_tag = {'gid': '649069647070258', 'resource_type': 'tag'}
    hold_tag =  {'gid': '259956811260129', 'resource_type': 'tag'}
    rownum = 11
    for row, task in enumerate(tasks):
        ##OPYXL is 1-indexed. This should be the first row of the output spreadsheet you want to fill. TODO: Find the actual header and drop down one instead of writing '11'
            print(task['name'])
            if hold_tag not in task['tags']:
                if qa_tag in task['tags']:
                    for ind, field in enumerate(task['custom_fields']):
                        print(field['name'])
                        ##Add review status...
                        if (field['name'] == 'Comment Tracking'):
                            sheet.cell(row = rownum, column = 2).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Submittal ID (CDRL number and section)'):
                            sheet.cell(row = rownum, column = 5).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'MBTA Owner'):
                            sheet.cell(row = rownum, column = 3).value = task['custom_fields'][ind]['display_value']
                        if (field['name'] == 'Design Review Comment Status'):
                            sheet.cell(row = rownum, column = 11).value = task['custom_fields'][ind]['display_value']
                    sheet.cell(row = rownum, column = 7).value = task['notes']
                    sheet.cell(row = rownum, column = 6).value = task['name']
                    sheet.cell(row = rownum, column = 4).value = task['section']
                    time = task['created_at']
                    split_time = time.split('T')[0]
                    strip_time = datetime.datetime.strptime(split_time, "%Y-%m-%d")
                    time_str = strip_time.strftime('%m/%d/%Y')
                    print(time_str)

                    sheet.cell(row = rownum, column = 8).value = time_str
                    rownum = rownum + 1
            
            workbook.save('output.xlsm')

def get_tasks_by_section(section_gid):
    ##Get section name to append to each task....
    section_name = client.sections.find_by_id(section_gid)['name']
    
    tasklist = []
    tasks = client.tasks.get_tasks({'section':section_gid, 'completed_since':date.today(), 'opt_fields':['name', 'notes', 'tags', 'custom_fields', 'created_at', 'memberships','permalink_url']}, )
    print(tasks)
    for task in tasks:
        print("Appending from " + section_name)
        updated_task = task
        updated_task['section'] = section_name
        tasklist.append(task)
    return tasklist

def display_project_name(project_gid):
    project = client.projects.get_project(project_gid)
    
    return(project['name'])
    
def display_section_name(section_gid):
    section = client.sections.get_section(section_gid)
    return(section['name'])

def define_workbook(mode):
    ##Get test template & compare to test input
    if mode == 'test':
        print('test workbook loading')
        template_contents = os.listdir('template/test.')
        if(len(template_contents)>1):
            print('!ERROR! More than one file in template test folder')
        else:
            tem_filename = os.listdir('template/test')[0]
            tem_filepath = os.path.dirname(os.path.abspath(__file__)) + '/template/test/'+ tem_filename
        temp_xlsx=pandas.ExcelFile(tem_filepath)
        ##Select the 'Document Comments' worksheet
        df_temp = pandas.read_excel(temp_xlsx, "Document Comments")

        input_contents = os.listdir('input/test/')
        if(len(template_contents)>1):
            print('!ERROR! More than one file in input folder')
        else:
            in_filename = os.listdir('input/test')[0]
            in_filepath = os.path.dirname(os.path.abspath(__file__)) + '/input/test/'+ in_filename
        in_xlsx=pandas.ExcelFile(in_filepath)
        ##Select the 'Document Comments' worksheet
        df_in = pandas.read_excel(in_xlsx, "Document Comments")
        ##Row 8 (pandas) contains the headers. Make sure these match.
        df_temp_headers = df_temp.iloc[8]
        df_in_headers = df_in.iloc[8]
        print(df_in_headers)



        if(df_in_headers.equals(df_temp_headers)):
             print("This program has determined that row 8 of both the input and template match.")
             in_opyxl_wb = openpyxl.load_workbook(in_filepath, keep_vba=True)
             in_opyxl_sheet = in_opyxl_wb['Document Comments']

             return {'workbook': in_opyxl_wb, 'worksheet': in_opyxl_sheet} 
        else:
            print("This program has determined that row 8 of the input and template are different. Does your input format match the template exactly?")
            quit()

    ##Standard mode
    if mode == 'standard':
        print('standard workbook loading')
        template_contents = os.listdir('template/standard.')
        if(len(template_contents)>1):
            print('!ERROR! More than one file in template test folder')
        else:
            tem_filename = os.listdir('template/standard')[0]
            tem_filepath = os.path.dirname(os.path.abspath(__file__)) + '/template/standard/'+ tem_filename
        temp_xlsx=pandas.ExcelFile(tem_filepath)
        ##Select the 'Document Comments' worksheet
        df_temp = pandas.read_excel(temp_xlsx, "Document Comments")

        input_contents = os.listdir('input/standard/')
        if(len(template_contents)>1):
            print('!ERROR! More than one file in input folder')
        else:
            in_filename = os.listdir('input/standard')[0]
            in_filepath = os.path.dirname(os.path.abspath(__file__)) + '/input/standard/'+ in_filename
        in_xlsx=pandas.ExcelFile(in_filepath)
        ##Select the 'Document Comments' worksheet
        df_in = pandas.read_excel(in_xlsx, "Document Comments")
        ##Row 8 (pandas) contains the headers. Make sure these match.
        df_temp_headers = df_temp.iloc[8]
        df_in_headers = df_in.iloc[8]
        print(df_in_headers)



        if(df_in_headers.equals(df_temp_headers)):
             print("This program has determined that row 8 of both the input and template match.")
             in_opyxl_wb = openpyxl.load_workbook(in_filepath, keep_vba=True)
             in_opyxl_sheet = in_opyxl_wb['Document Comments']

             return {'workbook': in_opyxl_wb, 'worksheet': in_opyxl_sheet} 
        else:
            print("This program has determined that row 8 of the input and template are different. Does your input format match the template exactly?")
            quit()

    


    ##Get standard template & compare to standard input
    if mode  == 'standard':
        print('stand')

def choose_sections_standard():
    
    sections = []
    project_choice = ""
    section_choice = ""

    print('Please enter a project GID. To quit, enter `quit`\n')
    while project_choice != "quit":
        project_choice = input('')
        if project_choice == 'quit':
            break
        project_sections = client.sections.find_by_project(project_choice)
        print('You are accessing ' + display_project_name(project_choice))
        print("Your available sections are")
        for el in project_sections:
            print(el)
        if project_choice != "":
            break
##Could make an index map to make this faster, but might not save this time. 
    
    print('Please enter a section gid. To quit, enter `quit`. To package, enter `package`\n')
    while section_choice != "quit":
        section_choice=input('')
        if section_choice == 'quit':
            break
        if section_choice == 'package':
            ##Pass section GIDs to function written here that displays names for confirmation, prompts for confirmation before writing to spreadsheets.
            print('Packaging function executing')
            tasklist = []
            for section in sections:
                tasklist.extend(get_tasks_by_section(section))
            opyxl = define_workbook('standard')
            AddStandardTasksToWorksheetNoLinks(tasklist, opyxl['worksheet'], opyxl['workbook'])
            break
        sections.append(section_choice)
        print("Currently selected sections:")
        for section in sections:
            print(section)
            print(display_section_name(section))

def choose_test_project():
    test_project_id = input("Please enter the ID of the test project (E.G: 'Test: System Website' is `1202257161854797` as extracted from https://app.asana.com/0/1202257161854797/list)")
    general_gid = None
    cases_gid = None
    test_project_sections = client.sections.find_by_project(test_project_id)
    for section in test_project_sections:
        print(section)
        if(section['name'] == 'General comments'):
            if(general_gid != None):
                print("WARNING! Multiple sections found with the name, 'General comments'. Quitting.")
            general_gid = section['gid']
        if(section['name'] == 'Test cases currently under review'):
            if(cases_gid != None):
                print("Warning! Multiple sections labeled 'Test cases under review'. Quitting.")
                quit()
            cases_gid = section['gid']
    gen_tasklist = get_tasks_by_section(general_gid)

    cases_tasklist = get_tasks_by_section(cases_gid)
    ##Note to self learn about these lambda key sorts these are slick
    cases_tasklist.sort(key=lambda x: x['name'])
    opyxl = define_workbook('test')
    AddTestTasksToWorksheet(gen_tasklist, cases_tasklist, opyxl['worksheet'], opyxl['workbook'])


##SIMPLE ASANA AUTHENTICATION##
##Headers to log in as Robert
##Put your token in a 'credentials.py' in the same directory as this script
##Set token equal to a Personal Access Token
sys.path.append(os.path.relpath('.\credentials.py'))
from credentials import token
client = asana.Client.access_token(token)
workspace = '15492006741476' ##MBTA Workspace

sample_task = client.tasks.find_by_id('1203367193397454')
print(sample_task)
print(sample_task.keys())

start_input = ""
print(logo_str)
print("Welcome to the Asana packager. Are you creating a test script package or a standard package? \n Enter 'test', 'standard' or 'quit'")
while start_input != "quit":
    start_input=input('')
    if start_input == "quit":
        break
    if start_input == "standard":
        choose_sections_standard()
        break
    if start_input == "test":
        choose_test_project()
        break

